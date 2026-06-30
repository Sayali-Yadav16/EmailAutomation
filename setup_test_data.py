#!/usr/bin/env python3
"""Create test patients.xlsx file for UiPath workflow"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'

# Add headers
headers = ['Name', 'Age', 'Email', 'Doctor', 'Appointment_Date', 'Slot', 'Priority']
ws.append(headers)

# Style headers
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add sample patient data
patients = [
    ['Rahul Sharma', 35, 'rahul.sharma@email.com', 'Cardiology', (datetime.now() + timedelta(days=3)).strftime('%Y-%m-%d'), '10:00 AM', 'High Priority'],
    ['Sayali Patel', 28, 'sayali.patel@email.com', 'Neurology', (datetime.now() + timedelta(days=5)).strftime('%Y-%m-%d'), '02:00 PM', 'Normal'],
    ['Arjun Singh', 45, 'arjun.singh@email.com', 'Orthopedics', (datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d'), '11:30 AM', 'High Priority'],
    ['Priya Verma', 32, 'priya.verma@email.com', 'Dermatology', (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d'), '03:30 PM', 'Normal'],
    ['Vikram Gupta', 55, 'vikram.gupta@email.com', 'Cardiology', (datetime.now() + timedelta(days=4)).strftime('%Y-%m-%d'), '01:00 PM', 'High Priority'],
    ['Neha Sharma', 29, 'neha.sharma@email.com', 'Gynecology', (datetime.now() + timedelta(days=6)).strftime('%Y-%m-%d'), '09:00 AM', 'Normal'],
    ['Amit Kumar', 52, 'amit.kumar@email.com', 'Urology', (datetime.now() + timedelta(days=8)).strftime('%Y-%m-%d'), '04:00 PM', 'Normal'],
    ['Divya Nair', 38, 'divya.nair@email.com', 'Pediatrics', (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d'), '10:30 AM', 'High Priority'],
]

for patient in patients:
    ws.append(patient)

# Set column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 15

# Save file
wb.save('patients.xlsx')
print('✅ patients.xlsx created successfully with 8 sample patients')
print('📊 File location: d:\\SEM 6\\RPA\\RPA Project\\patients.xlsx')
